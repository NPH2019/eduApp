from datetime import datetime

import openpyxl
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

# Create your views here.
from django.utils.crypto import get_random_string
from django.views.decorators.http import require_http_methods
from text_unidecode import unidecode

from eduApp.backend.client.forms import ClientCreateForm
from eduApp.backend.client.models import UserClient
from eduApp.backend.study_program.models import Lesson, Topic, Class


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def index(request):
    if request.method == 'GET':
        obj_client = UserClient.objects.all()
        return render(request, 'user_client/index.html', {
            'obj_client': obj_client
        })


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def create(request):
    if request.method == 'POST':
        forms = ClientCreateForm(request.POST)
        if forms.is_valid():
            if request.POST['amount']:
                max_row = request.POST['amount']
                for row in range(1, int(max_row) + 1):
                    obj = UserClient()
                    obj.client_name = get_random_string(length=8)
                    obj.client_password = get_random_string(length=8)
                    obj.client_code = request.POST['code']
                    obj.client_status = 0
                    obj.save()
        return redirect('client:index-client')

    return render(request, 'user_client/create.html')

#
# @login_required(login_url='login')
# @require_http_methods(["GET", "POST"])
# def tool(request):
#     wb_obj = openpyxl.load_workbook('media/yhoc_tai_ve.xlsx')
#     sheet_obj = wb_obj.active
#     max_row = sheet_obj.max_row
#     # Lesson.objects.filter(lesson_topic_id__in=[351, 352, 353, 354, 355, 356, 357, 358, 359, 360, 361]).delete()
#     # a = Class()
#     # a.class_name = "TỪ VƯNG TOEIC"
#     # a.class_program_id = 5
#     # a.save()
#     for row in range(1, max_row + 1):
#         ten_mon = sheet_obj.cell(row, 1).value
#         code = unidecode(ten_mon).replace(' ', '')
#         t = Topic.objects.get(topic_code=code, topic_class_id=44)
#         d = Lesson()
#         d.lesson_name = sheet_obj.cell(row, 2).value
#         d.lesson_url = sheet_obj.cell(row, 6).value
#         d.lesson_topic_id = t.topic_id
#         d.save()
#     # else:
#     # l = Topic()
#     # l.topic_name = "TỪ VƯNG TOEIC"
#     # l.topic_code = "TUVUNGTOEIC"
#     # l.topic_class_id = 37
#     # l.save()
#
#     #     d = Lesson()
#     #     d.lesson_name = sheet_obj.cell(row, 2).value
#     #     d.lesson_url = sheet_obj.cell(row, 3).value
#     #     d.lesson_topic_id = l.topic_id
#     #     d.save()
#
#     return render(request, 'user_client/index.html')
