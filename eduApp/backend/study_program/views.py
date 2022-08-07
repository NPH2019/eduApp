import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
from unidecode import unidecode

from .forms import ProgramCreateForm, ClassCreateForm, TopicCreateForm, LessonCreateForm
from .models import Program, Class, Topic, Lesson
from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods


# Class
@login_required(login_url='login')
@require_http_methods(["GET"])
def ad_class(request):
    if request.method == 'GET':
        obj = Class.objects.all()
        return render(request, 'class/index.html', {'obj': obj})


# Create class
@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def ad_class_create(request):
    obj_program = Program.objects.all()
    if request.method == 'POST':
        forms = ClassCreateForm(request.POST)
        if forms.is_valid():
            class_name = request.POST['class_name']
            class_detail = request.POST['class_detail']
            class_code = request.POST['class_code']

            obj = Class()
            if 'class_status' in request.POST:
                class_status = request.POST['class_status']
                if class_status != '':
                    class_status = True
            else:
                class_status = False
            if 'class_program' in request.POST:
                class_program = request.POST['class_program']
                if class_program != '':
                    obj.class_program_id = class_program

            obj.class_name = class_name
            obj.class_detail = class_detail
            obj.class_status = class_status
            obj.class_code = class_code
            obj.save()

            return redirect('study-program:class')
    return render(request, 'class/create.html', {'obj': obj_program})


# Lesson
@login_required(login_url='login')
@require_http_methods(["GET"])
def lesson(request):
    if request.method == 'GET':
        obj = Lesson.objects.all()
        return render(request, 'lesson/index.html', {'obj': obj})


# Create lesson
@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def lesson_create(request):
    if request.method == 'GET':
        obj_topic = Topic.objects.all()
        return render(request, 'lesson/create.html', {
            'obj': obj_topic,
        })
    else:
        forms = LessonCreateForm(request.POST, request.FILES)
        if forms.is_valid():
            lesson_name = request.POST['lesson_name']
            lesson_detail = request.POST['lesson_detail']
            lesson_code = request.POST['lesson_code']
            lesson_url = request.POST['lesson_url']

            lesson_files = ''
            if len(request.FILES) != 0:
                file_lesson = request.FILES['lesson_files']
                lesson_files = file_lesson.name
                fs = FileSystemStorage()
                filename = fs.save(lesson_files, file_lesson)
                lesson_files = fs.url(filename)

            obj = Lesson()
            if 'lesson_status' in request.POST:
                lesson_status = request.POST['lesson_status']
                if lesson_status != '':
                    lesson_status = True
            else:
                lesson_status = False
            if 'lesson_topic' in request.POST:
                lesson_topic = request.POST['lesson_topic']
                if lesson_topic != '':
                    obj.lesson_topic_id = lesson_topic

            obj.lesson_name = lesson_name
            obj.lesson_files = lesson_files
            obj.lesson_detail = lesson_detail
            obj.lesson_status = lesson_status
            obj.lesson_code = lesson_code
            obj.lesson_url = lesson_url
            obj.save()

            return redirect('study-program:lesson')


# Topic
@login_required(login_url='login')
@require_http_methods(["GET"])
def topic(request):
    if request.method == 'GET':
        obj = Topic.objects.all()
        return render(request, 'topic/index.html', {'obj': obj})


# Create topic
@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def topic_create(request):
    obj_class = Class.objects.all()
    if request.method == 'POST':
        forms = TopicCreateForm(request.POST)
        if forms.is_valid():
            topic_name = request.POST['topic_name']
            topic_detail = request.POST['topic_detail']
            topic_code = request.POST['topic_code']

            obj = Topic()
            if 'topic_status' in request.POST:
                topic_status = request.POST['topic_status']
                if topic_status != '':
                    topic_status = True
            else:
                topic_status = False
            if 'topic_class' in request.POST:
                topic_class = request.POST['topic_class']
                if topic_class != '':
                    obj.topic_class_id = topic_class

            obj.topic_name = topic_name
            obj.topic_detail = topic_detail
            obj.topic_status = topic_status
            obj.topic_code = topic_code
            obj.save()

            return redirect('study-program:topic')
    return render(request, 'topic/create.html', {
        'obj': obj_class,
    })


# Programme
@login_required(login_url='login')
@require_http_methods(["GET"])
def program(request):
    if request.method == 'GET':
        obj = Program.objects.all()
        return render(request, 'program/index.html', {'obj': obj})


# Programme create
@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def program_create(request):
    if request.method == 'POST':
        forms = ProgramCreateForm(request.POST)
        if forms.is_valid():
            program_name = request.POST['program_name']
            program_detail = request.POST['program_detail']
            program_code = request.POST['program_code']

            if 'program_status' in request.POST:
                program_status = request.POST['program_status']
                if program_status != '':
                    program_status = True
            else:
                program_status = False
            obj = Program()
            obj.program_name = program_name
            obj.program_code = program_code
            obj.program_detail = program_detail
            obj.program_status = program_status
            obj.save()
            return redirect('study-program:program')

    return render(request, 'program/create.html')


@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def client(request):
    # wb_obj = openpyxl.load_workbook('media/1_tai_ve.xlsx')
    # sheet_obj = wb_obj.active
    # max_row = sheet_obj.max_row
    # Lesson.objects.filter(lesson_topic_id=197).delete()
    # a = Class()
    # a.class_name = "Y Học"
    # a.class_program_id = 5
    # a.save()
    # for row in range(1, max_row + 1):
    #     ten_mon = sheet_obj.cell(row, 1).value
    #     code = unidecode(ten_mon).replace(' ', '')
    #     f = Topic.objects.filter(topic_code=code, topic_class_id=25).count()
    #     if f > 0:
    #         t = Topic.objects.get(topic_code=code, topic_class_id=25)
    #         d = Lesson()
    #         d.lesson_name = sheet_obj.cell(row, 2).value
    #         d.lesson_url = sheet_obj.cell(row, 3).value
    #         d.lesson_topic_id = t.topic_id
    #         d.save()
    #     else:
    #         l = Topic()
    #         l.topic_name = ten_mon
    #         l.topic_code = code
    #         l.topic_class_id = 25
    #         l.save()
    #         d = Lesson()
    #         d.lesson_name = sheet_obj.cell(row, 2).value
    #         d.lesson_url = sheet_obj.cell(row, 3).value
    #         d.lesson_topic_id = l.topic_id
    #         d.save()

    return render(request, 'user_client/index.html')
